#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel İşleme Uygulaması - Rapor Modülü (Düzeltilmiş)
Analiz sonuçları için Excel/PDF rapor oluşturma sistemi
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import pandas as pd
from utils import format_turkish_number
import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self):
        self.report_templates = {
            'arac_summary': 'ARAÇ Özet Raporu',
            'assignment_report': 'Sorumlu Atama Raporu',
            'detailed_analysis': 'Detaylı Analiz Raporu',
            'comparison_report': 'ARAÇ Karşılaştırma Raporu',
            'aging_analysis': 'Yaşlandırma Analizi Raporu'
        }
    
    def generate_arac_summary_report(self, analysis_results: Dict, assignments: Dict = None) -> Optional[pd.DataFrame]:
        """ARAÇ özet raporu oluştur - DÜZELTİLMİŞ"""
        try:
            if not analysis_results:
                logger.error("Rapor için analiz verisi bulunamadı")
                return None
            
            report_data = []
            
            for arac_no, analysis in analysis_results.items():
                # Atama bilgisi
                sorumlu = "Atanmamış"
                if assignments and arac_no in assignments:
                    sorumlu = assignments[arac_no].get('sorumlu', 'Atanmamış')
                
                # İstatistikler
                stats = analysis.get('istatistikler', {})
                
                row_data = {
                    'ARAÇ No': arac_no,
                    'Sorumlu': sorumlu,
                    'Müşteri Sayısı': analysis.get('musteri_sayisi', 0),
                    'Toplam Bakiye (TL)': analysis.get('toplam_bakiye', 0),
                    'Toplam Bakiye (Formatlanmış)': format_turkish_number(analysis.get('toplam_bakiye', 0)),
                    'Açık Hesap (TL)': analysis.get('acik_hesap', 0),
                    'Açık Hesap (Formatlanmış)': format_turkish_number(analysis.get('acik_hesap', 0)),
                    'Ortalama Bakiye (TL)': stats.get('ortalama_bakiye', 0),
                    'Ortalama Bakiye (Formatlanmış)': format_turkish_number(stats.get('ortalama_bakiye', 0)),
                    'En Yüksek Bakiye': format_turkish_number(stats.get('en_yuksek_bakiye', 0)),
                    'En Düşük Bakiye': format_turkish_number(stats.get('en_dusuk_bakiye', 0)),
                    'Pozitif Bakiye Sayısı': stats.get('bakiye_pozitif_olan', 0),
                    'Negatif Bakiye Sayısı': stats.get('bakiye_negatif_olan', 0),
                    'Analiz Tarihi': analysis.get('analiz_tarihi', '')
                }
                
                report_data.append(row_data)
            
            # DataFrame oluştur
            df = pd.DataFrame(report_data)
            
            # Sıralama (ARAÇ numarasına göre)
            df = df.sort_values('ARAÇ No')
            
            logger.info(f"ARAÇ özet raporu oluşturuldu: {len(df)} kayıt")
            return df
            
        except Exception as e:
            logger.error(f"ARAÇ özet raporu oluşturma hatası: {e}")
            return None
    
    def generate_detailed_analysis_report(self, analysis_results: Dict) -> Optional[pd.DataFrame]:
        """Detaylı analiz raporu oluştur - DÜZELTİLMİŞ"""
        try:
            if not analysis_results:
                logger.error("Detaylı rapor için analiz verisi bulunamadı")
                return None
            
            report_data = []
            
            for arac_no, analysis in analysis_results.items():
                musteri_detaylari = analysis.get('musteri_detaylari', [])
                
                for musteri in musteri_detaylari:
                    row_data = {
                        'ARAÇ No': arac_no,
                        'Cari Ünvan': musteri.get('cari_unvan', ''),
                        'Toplam Bakiye (TL)': musteri.get('toplam_bakiye', 0),
                        'Toplam Bakiye (Formatlanmış)': format_turkish_number(musteri.get('toplam_bakiye', 0))
                    }
                    
                    # Yaşlandırma detayları
                    bakiye_detay = musteri.get('bakiye_detay', {})
                    for period, amount in bakiye_detay.items():
                        row_data[f'{period} (TL)'] = amount
                        row_data[f'{period} (Formatlanmış)'] = format_turkish_number(amount)
                    
                    report_data.append(row_data)
            
            if not report_data:
                logger.warning("Detaylı rapor için veri bulunamadı")
                return None
            
            # DataFrame oluştur
            df = pd.DataFrame(report_data)
            
            # Sıralama
            df = df.sort_values(['ARAÇ No', 'Cari Ünvan'])
            
            logger.info(f"Detaylı analiz raporu oluşturuldu: {len(df)} kayıt")
            return df
            
        except Exception as e:
            logger.error(f"Detaylı analiz raporu oluşturma hatası: {e}")
            return None
    
    def generate_assignment_report(self, assignments: Dict) -> Optional[pd.DataFrame]:
        """Sorumlu atama raporu oluştur"""
        try:
            if not assignments:
                logger.error("Atama raporu için veri bulunamadı")
                return None
            
            report_data = []
            
            for arac_no, assignment in assignments.items():
                row_data = {
                    'ARAÇ No': arac_no,
                    'Sorumlu': assignment.get('sorumlu', ''),
                    'Email': assignment.get('email', ''),
                    'Telefon': assignment.get('telefon', ''),
                    'Departman': assignment.get('departman', ''),
                    'Atama Tarihi': assignment.get('atama_tarihi', ''),
                    'Durum': assignment.get('durum', ''),
                    'Notlar': assignment.get('notlar', ''),
                    'Son Güncelleme': assignment.get('son_guncelleme', '')
                }
                
                report_data.append(row_data)
            
            # DataFrame oluştur
            df = pd.DataFrame(report_data)
            
            # Sıralama
            df = df.sort_values(['Sorumlu', 'ARAÇ No'])
            
            logger.info(f"Atama raporu oluşturuldu: {len(df)} kayıt")
            return df
            
        except Exception as e:
            logger.error(f"Atama raporu oluşturma hatası: {e}")
            return None
    
    def generate_aging_analysis_report(self, analysis_results: Dict) -> Optional[pd.DataFrame]:
        """Yaşlandırma analizi raporu oluştur"""
        try:
            if not analysis_results:
                logger.error("Yaşlandırma raporu için analiz verisi bulunamadı")
                return None
            
            report_data = []
            
            for arac_no, analysis in analysis_results.items():
                yaslanding_analizi = analysis.get('yaslanding_analizi', {})
                
                row_data = {
                    'ARAÇ No': arac_no,
                    'Toplam Bakiye (TL)': analysis.get('toplam_bakiye', 0),
                    'Toplam Bakiye (Formatlanmış)': format_turkish_number(analysis.get('toplam_bakiye', 0))
                }
                
                # Her yaşlandırma dönemi için
                for period, amount in yaslanding_analizi.items():
                    row_data[f'{period} (TL)'] = amount
                    row_data[f'{period} (Formatlanmış)'] = format_turkish_number(amount)
                    
                    # Yüzde hesapla
                    total_balance = analysis.get('toplam_bakiye', 0)
                    if total_balance > 0:
                        percentage = (amount / total_balance) * 100
                        row_data[f"{period} (%)"] = f"{percentage:.1f}%"
                    else:
                        row_data[f"{period} (%)"] = "0.0%"
                
                report_data.append(row_data)
            
            if not report_data:
                logger.warning("Yaşlandırma raporu için veri bulunamadı")
                return None
            
            # DataFrame oluştur
            df = pd.DataFrame(report_data)
            
            # Sıralama
            df = df.sort_values('ARAÇ No')
            
            logger.info(f"Yaşlandırma analizi raporu oluşturuldu: {len(df)} kayıt")
            return df
            
        except Exception as e:
            logger.error(f"Yaşlandırma analizi raporu oluşturma hatası: {e}")
            return None
    
    def generate_comparison_report(self, analysis_results: Dict, arac_list: List[str] = None) -> Optional[pd.DataFrame]:
        """ARAÇ karşılaştırma raporu oluştur"""
        try:
            if not analysis_results:
                logger.error("Karşılaştırma raporu için analiz verisi bulunamadı")
                return None
            
            # ARAÇ listesi belirtilmemişse tümünü al
            if not arac_list:
                arac_list = list(analysis_results.keys())
            
            report_data = []
            
            for arac_no in arac_list:
                if arac_no not in analysis_results:
                    continue
                
                analysis = analysis_results[arac_no]
                stats = analysis.get('istatistikler', {})
                
                row_data = {
                    'ARAÇ No': arac_no,
                    'Müşteri Sayısı': analysis.get('musteri_sayisi', 0),
                    'Toplam Bakiye (TL)': analysis.get('toplam_bakiye', 0),
                    'Toplam Bakiye (Formatlanmış)': format_turkish_number(analysis.get('toplam_bakiye', 0)),
                    'Açık Hesap (TL)': analysis.get('acik_hesap', 0),
                    'Açık Hesap (Formatlanmış)': format_turkish_number(analysis.get('acik_hesap', 0)),
                    'Ortalama Bakiye (TL)': stats.get('ortalama_bakiye', 0),
                    'Ortalama Bakiye (Formatlanmış)': format_turkish_number(stats.get('ortalama_bakiye', 0)),
                    'En Yüksek Bakiye': format_turkish_number(stats.get('en_yuksek_bakiye', 0)),
                    'Pozitif Bakiye Oranı': f"{(stats.get('bakiye_pozitif_olan', 0) / analysis.get('musteri_sayisi', 1) * 100):.1f}%" if analysis.get('musteri_sayisi', 0) > 0 else "0.0%"
                }
                
                report_data.append(row_data)
            
            if not report_data:
                logger.warning("Karşılaştırma raporu için veri bulunamadı")
                return None
            
            # DataFrame oluştur
            df = pd.DataFrame(report_data)
            
            # Sıralama (bakiyeye göre)
            df_sorted_by_balance = df.sort_values('Toplam Bakiye (TL)', ascending=False)
            
            logger.info(f"Karşılaştırma raporu oluşturuldu: {len(df)} ARAÇ")
            return df_sorted_by_balance
            
        except Exception as e:
            logger.error(f"Karşılaştırma raporu oluşturma hatası: {e}")
            return None
    
    def save_report_to_excel(self, dataframe: pd.DataFrame, file_path: str, sheet_name: str = "Rapor") -> bool:
        """Raporu Excel dosyasına kaydet - İYİLEŞTİRİLMİŞ"""
        try:
            if dataframe is None or dataframe.empty:
                logger.error("Kaydedilecek rapor verisi yok")
                return False
            
            # Excel writer oluştur
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Ana raporu kaydet
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Worksheet'i al ve formatla
                worksheet = writer.sheets[sheet_name]
                
                # Sütun genişliklerini ayarla
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Maksimum 50 karakter
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Başlık satırını formatla
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                header_font = Font(bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
            
            logger.info(f"Rapor Excel'e kaydedildi: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Excel rapor kaydetme hatası: {e}")
            return False
    
    def save_multiple_reports_to_excel(self, reports: Dict[str, pd.DataFrame], file_path: str) -> bool:
        """Birden fazla raporu tek Excel dosyasına kaydet"""
        try:
            if not reports:
                logger.error("Kaydedilecek rapor yok")
                return False
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, dataframe in reports.items():
                    if dataframe is not None and not dataframe.empty:
                        # Sayfa adını kısalt (Excel limit 31 karakter)
                        safe_sheet_name = sheet_name[:31]
                        
                        dataframe.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                        
                        # Formatla
                        worksheet = writer.sheets[safe_sheet_name]
                        
                        # Sütun genişlikleri
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Başlık formatı
                        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        header_font = Font(bold=True)
                        
                        for cell in worksheet[1]:
                            cell.fill = header_fill
                            cell.font = header_font
            
            logger.info(f"Çoklu rapor Excel'e kaydedildi: {file_path} ({len(reports)} sayfa)")
            return True
            
        except Exception as e:
            logger.error(f"Çoklu Excel rapor kaydetme hatası: {e}")
            return False
    
    def generate_summary_statistics(self, analysis_results: Dict) -> Dict:
        """Özet istatistikler oluştur"""
        try:
            if not analysis_results:
                return {}
            
            total_aracs = len(analysis_results)
            total_customers = sum(analysis.get('musteri_sayisi', 0) for analysis in analysis_results.values())
            total_balance = sum(analysis.get('toplam_bakiye', 0) for analysis in analysis_results.values())
            total_open_balance = sum(analysis.get('acik_hesap', 0) for analysis in analysis_results.values())
            
            # En yüksek/düşük değerler
            balances = [analysis.get('toplam_bakiye', 0) for analysis in analysis_results.values()]
            customer_counts = [analysis.get('musteri_sayisi', 0) for analysis in analysis_results.values()]
            
            summary = {
                'genel_ozet': {
                    'toplam_arac_sayisi': total_aracs,
                    'toplam_musteri_sayisi': total_customers,
                    'toplam_bakiye': total_balance,
                    'toplam_acik_hesap': total_open_balance,
                    'ortalama_musteri_per_arac': total_customers / total_aracs if total_aracs > 0 else 0,
                    'ortalama_bakiye_per_arac': total_balance / total_aracs if total_aracs > 0 else 0
                },
                'arac_dagilimi': {
                    'en_yuksek_bakiye': max(balances) if balances else 0,
                    'en_dusuk_bakiye': min(balances) if balances else 0,
                    'en_fazla_musteri': max(customer_counts) if customer_counts else 0,
                    'en_az_musteri': min(customer_counts) if customer_counts else 0
                },
                'yaslanding_ozeti': self._calculate_aging_summary(analysis_results),
                'rapor_tarihi': datetime.now().isoformat()
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Özet istatistik oluşturma hatası: {e}")
            return {}
    
    def _calculate_aging_summary(self, analysis_results: Dict) -> Dict:
        """Yaşlandırma özeti hesapla"""
        try:
            aging_summary = {}
            total_balance = 0
            
            for analysis in analysis_results.values():
                yaslanding_analizi = analysis.get('yaslanding_analizi', {})
                
                for period, amount in yaslanding_analizi.items():
                    if period not in aging_summary:
                        aging_summary[period] = 0
                    aging_summary[period] += amount
                    total_balance += amount
            
            # Yüzde hesapla
            aging_percentages = {}
            for period, amount in aging_summary.items():
                if total_balance > 0:
                    aging_percentages[f"{period}_yuzde"] = (amount / total_balance) * 100
                else:
                    aging_percentages[f"{period}_yuzde"] = 0
            
            aging_summary.update(aging_percentages)
            aging_summary['toplam_bakiye'] = total_balance
            
            return aging_summary
            
        except Exception as e:
            logger.error(f"Yaşlandırma özeti hesaplama hatası: {e}")
            return {}
    
    def create_dashboard_data(self, analysis_results: Dict, assignments: Dict = None) -> Dict:
        """Dashboard için veri hazırla"""
        try:
            if not analysis_results:
                return {}
            
            # Temel istatistikler
            summary_stats = self.generate_summary_statistics(analysis_results)
            
            # ARAÇ performans verileri
            arac_performance = []
            for arac_no, analysis in analysis_results.items():
                performance_data = {
                    'arac_no': arac_no,
                    'musteri_sayisi': analysis.get('musteri_sayisi', 0),
                    'toplam_bakiye': analysis.get('toplam_bakiye', 0),
                    'acik_hesap': analysis.get('acik_hesap', 0),
                    'sorumlu': assignments.get(arac_no, {}).get('sorumlu', 'Atanmamış') if assignments else 'Atanmamış'
                }
                arac_performance.append(performance_data)
            
            # En iyi/kötü performans
            arac_performance.sort(key=lambda x: x['toplam_bakiye'], reverse=True)
            top_performers = arac_performance[:5]  # En iyi 5
            bottom_performers = arac_performance[-5:]  # En kötü 5
            
            # Yaşlandırma grafiği için veri
            aging_chart_data = []
            aging_summary = summary_stats.get('yaslanding_ozeti', {})
            
            for period, amount in aging_summary.items():
                if not period.endswith('_yuzde') and period != 'toplam_bakiye':
                    aging_chart_data.append({
                        'period': period,
                        'amount': amount,
                        'percentage': aging_summary.get(f"{period}_yuzde", 0)
                    })
            
            dashboard_data = {
                'summary_statistics': summary_stats,
                'arac_performance': arac_performance,
                'top_performers': top_performers,
                'bottom_performers': bottom_performers,
                'aging_chart_data': aging_chart_data,
                'last_updated': datetime.now().isoformat()
            }
            
            return dashboard_data
            
        except Exception as e:
            logger.error(f"Dashboard veri hazırlama hatası: {e}")
            return {}
    
    def get_available_reports(self) -> Dict[str, str]:
        """Mevcut rapor türlerini getir"""
        return self.report_templates.copy()